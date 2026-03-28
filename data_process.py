from __future__ import annotations

import argparse
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass
from pathlib import Path

import polars as pl
from openpyxl import load_workbook

from app.core.id_cards import is_valid_id_no, normalize_id_no


DEFAULT_INPUT_DIR = Path("/Users/wanghao/workspace/Project/myself_code/shuiwu/新建文件夹")
DEFAULT_OUTPUT_CSV = Path("/Users/wanghao/workspace/Project/myself_code/shuiwu/新建文件夹/import_ready_merged.csv")
DEFAULT_CHUNK_ROWS = 2_000_000


@dataclass(frozen=True)
class ChunkWriteResult:
    path: Path
    rows_before_dedup: int
    rows_written: int
    duplicate_rows: int


def _norm_header(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip().lower().replace("_", "").replace(" ", "")


def _resolve_column(columns: list[str], aliases: tuple[str, ...], fallback_idx: int) -> str | None:
    alias_set = {_norm_header(alias) for alias in aliases}
    for column in columns:
        if _norm_header(column) in alias_set:
            return column
    if 0 <= fallback_idx < len(columns):
        return columns[fallback_idx]
    return None


def _empty_frame() -> pl.DataFrame:
    return pl.DataFrame(schema={"name": pl.String, "id_no": pl.String, "birth_year": pl.UInt16})


def _prepare_frame(df: pl.DataFrame) -> tuple[pl.DataFrame, int, int]:
    columns = [str(col) for col in df.columns]
    if not columns:
        return _empty_frame(), 0, 0

    name_col = _resolve_column(columns, ("姓名", "name"), 0)
    id_col = _resolve_column(columns, ("身份证号", "身份证", "证件号", "idno", "id_no"), 1)
    year_col = _resolve_column(columns, ("年份", "year", "年", "birth_year"), 2)

    selected = df.select(
        [
            pl.col(name_col).alias("name") if name_col else pl.lit(None).alias("name"),
            pl.col(id_col).alias("id_no") if id_col else pl.lit(None).alias("id_no"),
            pl.col(year_col).alias("birth_year") if year_col else pl.lit(None).alias("birth_year"),
        ]
    ).with_columns(
        [
            pl.col("name").cast(pl.Utf8, strict=False).fill_null("").str.strip_chars(),
            pl.col("id_no").cast(pl.Utf8, strict=False).fill_null("").map_elements(normalize_id_no, return_dtype=pl.String),
            pl.col("birth_year")
            .cast(pl.Utf8, strict=False)
            .fill_null("")
            .str.strip_chars()
            .str.replace_all(r"\.0+$", ""),
        ]
    )

    non_blank = selected.filter(
        ~((pl.col("name") == "") & (pl.col("id_no") == "") & (pl.col("birth_year") == ""))
    )
    total_rows = int(non_blank.height)
    if total_rows == 0:
        return _empty_frame(), 0, 0

    required = non_blank.filter(
        (pl.col("name") != "") & (pl.col("id_no") != "") & (pl.col("birth_year").str.contains(r"^\d{4}$"))
    )
    deduped = required.unique(subset=["id_no"], keep="first", maintain_order=True)
    if deduped.height == 0:
        return _empty_frame(), total_rows, total_rows

    valid_mask = pl.Series(
        name="_id_valid",
        values=[is_valid_id_no(value) for value in deduped.get_column("id_no").to_list()],
        dtype=pl.Boolean,
    )
    valid = deduped.with_columns(valid_mask).filter(pl.col("_id_valid")).drop("_id_valid")
    if valid.height == 0:
        return _empty_frame(), total_rows, total_rows

    valid = valid.with_columns(pl.col("birth_year").cast(pl.UInt16, strict=False))
    skipped_rows = total_rows - int(valid.height)
    return valid.select(["name", "id_no", "birth_year"]), total_rows, skipped_rows


def _load_excel_sheet(file_path: Path, sheet_name: str, engine: str) -> pl.DataFrame:
    return pl.read_excel(file_path, sheet_name=sheet_name, engine=engine)


def _load_csv(file_path: Path) -> pl.DataFrame:
    try:
        return pl.read_csv(file_path, try_parse_dates=False, infer_schema_length=None)
    except Exception:
        return pl.read_csv(file_path, encoding="gb18030", try_parse_dates=False, infer_schema_length=None)


def _iter_prepared_frames(file_path: Path, *, sheet_workers: int, excel_engine: str):
    if file_path.suffix.lower() == ".csv":
        frame, total_rows, skipped_rows = _prepare_frame(_load_csv(file_path))
        print(
            f"📄 读取 CSV：{file_path.name}，原始有效行 {total_rows:,}，"
            f"清洗后 {frame.height:,}，跳过 {skipped_rows:,}"
        )
        yield frame, total_rows, skipped_rows
        return

    wb = load_workbook(file_path, read_only=True, data_only=True)
    try:
        sheet_names = list(wb.sheetnames)
    finally:
        wb.close()

    worker_count = min(max(1, sheet_workers), max(1, len(sheet_names)))
    with ThreadPoolExecutor(max_workers=worker_count, thread_name_prefix="xlsx-sheet") as executor:
        futures = {
            executor.submit(_load_excel_sheet, file_path, sheet_name, excel_engine): sheet_name for sheet_name in sheet_names
        }
        for future in as_completed(futures):
            sheet_name = futures[future]
            df = future.result()
            frame, total_rows, skipped_rows = _prepare_frame(df)
            print(
                f"  └─ Sheet {sheet_name}：原始有效行 {total_rows:,}，"
                f"清洗后 {frame.height:,}，跳过 {skipped_rows:,}"
            )
            yield frame, total_rows, skipped_rows


def _cleanup_old_outputs(output_csv_path: Path, *, split_output: bool) -> None:
    output_csv_path.parent.mkdir(parents=True, exist_ok=True)
    if split_output:
        suffix = output_csv_path.suffix or ".csv"
        pattern = f"{output_csv_path.stem}.part*{suffix}"
        for old_file in output_csv_path.parent.glob(pattern):
            old_file.unlink(missing_ok=True)
        return
    output_csv_path.unlink(missing_ok=True)


def _build_output_path(output_csv_path: Path, *, chunk_index: int, split_output: bool) -> Path:
    if not split_output:
        return output_csv_path
    suffix = output_csv_path.suffix or ".csv"
    return output_csv_path.with_name(f"{output_csv_path.stem}.part{chunk_index:04d}{suffix}")


def _flush_buffer(
    frames: list[pl.DataFrame],
    *,
    output_csv_path: Path,
    chunk_index: int,
    sort_rows: bool,
    split_output: bool,
) -> tuple[list[pl.DataFrame], int, ChunkWriteResult | None]:
    if not frames:
        return [], chunk_index, None

    merged = frames[0] if len(frames) == 1 else pl.concat(frames, how="vertical_relaxed")
    rows_before_dedup = int(merged.height)
    merged = merged.unique(subset=["id_no"], keep="first", maintain_order=True)
    duplicate_rows = rows_before_dedup - int(merged.height)

    if sort_rows and merged.height > 1:
        merged = merged.sort(["birth_year", "name", "id_no"])

    next_chunk_index = chunk_index + 1
    output_path = _build_output_path(output_csv_path, chunk_index=next_chunk_index, split_output=split_output)
    merged.write_csv(output_path, include_header=True)

    result = ChunkWriteResult(
        path=output_path,
        rows_before_dedup=rows_before_dedup,
        rows_written=int(merged.height),
        duplicate_rows=duplicate_rows,
    )
    print(
        f"📦 写出分片 {next_chunk_index}：{output_path.name}，"
        f"分片行数 {result.rows_written:,}，分片内去重 {result.duplicate_rows:,}"
    )
    return [], next_chunk_index, result


def process_all_data(
    input_dir: Path,
    output_csv_path: Path,
    *,
    sheet_workers: int,
    excel_engine: str,
    chunk_rows: int,
    sort_rows: bool,
) -> None:
    split_output = chunk_rows > 0
    _cleanup_old_outputs(output_csv_path, split_output=split_output)

    input_files = sorted(
        [
            path
            for path in input_dir.iterdir()
            if path.is_file() and not path.name.startswith(".") and path.suffix.lower() in {".xlsx", ".csv"}
        ]
    )
    if not input_files:
        raise FileNotFoundError(f"在目录 {input_dir} 中未找到 .xlsx/.csv 文件")

    print(f"✅ 找到 {len(input_files)} 个输入文件")
    print("📌 输出结构固定为：name,id_no,birth_year")
    if split_output:
        print(f"📌 按分片输出，目标每片约 {chunk_rows:,} 行")
    else:
        print("📌 输出为单个 CSV 文件")

    total_rows = 0
    skipped_rows = 0
    kept_rows = 0
    chunk_index = 0
    chunk_results: list[ChunkWriteResult] = []
    buffer_frames: list[pl.DataFrame] = []
    buffer_rows = 0

    for file_idx, file_path in enumerate(input_files, start=1):
        print(f"📖 读取文件 {file_idx}/{len(input_files)}：{file_path.name}")
        for frame, frame_total_rows, frame_skipped_rows in _iter_prepared_frames(
            file_path,
            sheet_workers=sheet_workers,
            excel_engine=excel_engine,
        ):
            total_rows += int(frame_total_rows)
            skipped_rows += int(frame_skipped_rows)
            if frame.height == 0:
                continue

            kept_rows += int(frame.height)
            buffer_frames.append(frame)
            buffer_rows += int(frame.height)

            if split_output and buffer_rows >= chunk_rows:
                buffer_frames, chunk_index, result = _flush_buffer(
                    buffer_frames,
                    output_csv_path=output_csv_path,
                    chunk_index=chunk_index,
                    sort_rows=sort_rows,
                    split_output=split_output,
                )
                buffer_rows = 0
                if result is not None:
                    chunk_results.append(result)

    if buffer_frames:
        buffer_frames, chunk_index, result = _flush_buffer(
            buffer_frames,
            output_csv_path=output_csv_path,
            chunk_index=chunk_index,
            sort_rows=sort_rows,
            split_output=split_output,
        )
        if result is not None:
            chunk_results.append(result)

    total_written = sum(item.rows_written for item in chunk_results)
    total_chunk_duplicates = sum(item.duplicate_rows for item in chunk_results)

    print("🎉 全部处理完成")
    print(f"📊 原始有效行数：{total_rows:,}")
    print(f"📊 清洗后待输出行数：{kept_rows:,}")
    print(f"📊 清洗阶段跳过行数：{skipped_rows:,}")
    print(f"📊 分片内额外去重行数：{total_chunk_duplicates:,}")
    print(f"📊 最终写出总行数：{total_written:,}")
    if chunk_results:
        print(f"📁 第一片：{chunk_results[0].path}")
        print(f"📁 最后一片：{chunk_results[-1].path}")
    print("💡 导入时请直接使用这些 CSV 文件的服务器本地路径，不要再预加密身份证。")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="生成可直接导入 data_manager 的明文 CSV/分片 CSV")
    parser.add_argument("--input-dir", type=Path, default=DEFAULT_INPUT_DIR, help="包含原始 Excel/CSV 的目录")
    parser.add_argument(
        "--output-csv",
        type=Path,
        default=DEFAULT_OUTPUT_CSV,
        help="输出 CSV 基础路径；分片模式下会生成 .part0001.csv 这类文件",
    )
    parser.add_argument("--sheet-workers", type=int, default=4, help="Excel Sheet 并行读取线程数")
    parser.add_argument("--excel-engine", default="calamine", help="Polars Excel 引擎，默认 calamine")
    parser.add_argument(
        "--chunk-rows",
        type=int,
        default=DEFAULT_CHUNK_ROWS,
        help="每片目标行数；设为 0 表示输出单个 CSV",
    )
    parser.add_argument("--no-sort", action="store_true", help="不按 birth_year,name,id_no 排序")
    return parser.parse_args()


if __name__ == "__main__":
    args = parse_args()
    process_all_data(
        input_dir=args.input_dir,
        output_csv_path=args.output_csv,
        sheet_workers=max(1, int(args.sheet_workers)),
        excel_engine=str(args.excel_engine).strip() or "calamine",
        chunk_rows=max(0, int(args.chunk_rows)),
        sort_rows=not args.no_sort,
    )
