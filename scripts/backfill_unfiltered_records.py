from __future__ import annotations

import argparse
import os
from pathlib import Path

import polars as pl
from openpyxl import load_workbook

os.environ.setdefault("CLICKHOUSE_TIMEOUT_SECONDS", "7200")

from app.core.config import get_settings
from app.core.crypto import encrypt_id_values
from app.core.id_cards import fingerprint_id_no
from app.core.key_manager import load_keys
from app.db.clickhouse import clickhouse_command, clickhouse_insert_json_rows, clickhouse_query_rows
from app.services.clickhouse_records import ensure_clickhouse_record_store
from app.services.importer import _find_optional_polars_column, _resolve_polars_column


DEFAULT_INPUT_FILES = [
    Path("/Users/wanghao/workspace/Project/myself_code/shuiwu/新建文件夹/老人数据.xlsx"),
    Path("/Users/wanghao/workspace/Project/myself_code/shuiwu/新建文件夹/小孩数据.xlsx"),
]
DEFAULT_STAGE_TABLE = "person_records_backfill_stage"
DEFAULT_BATCH_SIZE = 100_000


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="回填 Excel 中此前被过滤/去重丢掉的记录")
    parser.add_argument(
        "--input-file",
        dest="input_files",
        action="append",
        type=Path,
        default=None,
        help="待回填的 Excel/CSV 文件；可传多次，默认使用老人数据.xlsx 和小孩数据.xlsx",
    )
    parser.add_argument("--stage-table", default=DEFAULT_STAGE_TABLE, help="临时 stage 表名")
    parser.add_argument("--batch-size", type=int, default=DEFAULT_BATCH_SIZE, help="每批写入 ClickHouse 的行数")
    parser.add_argument("--created-by", type=int, default=1, help="写入 created_by 字段的用户 ID")
    parser.add_argument("--keep-stage", action="store_true", help="回填后保留临时 stage 表")
    parser.add_argument("--use-existing-stage", action="store_true", help="复用已存在的临时 stage 表，不重复加载源文件")
    return parser.parse_args()


def get_sheet_names(file_path: Path) -> list[str]:
    if file_path.suffix.lower() == ".csv":
        return [""]
    wb = load_workbook(file_path, read_only=True, data_only=True)
    try:
        return list(wb.sheetnames)
    finally:
        wb.close()


def prepare_stage_frame(df: pl.DataFrame) -> pl.DataFrame:
    columns = [str(col) for col in df.columns]
    name_col = _resolve_polars_column(columns, ("姓名", "name"), 0)
    id_col = _resolve_polars_column(columns, ("身份证号", "身份证", "证件号", "idno", "id_no"), 1)
    year_col = _resolve_polars_column(columns, ("年份", "year", "年", "birth_year"), 2)
    year_raw_col = _find_optional_polars_column(columns, ("birth_year_raw", "yearraw", "年份原始", "原始年份"))

    indexed = df.with_row_index("source_row_no", offset=2)
    return indexed.select(
        [
            pl.col("source_row_no"),
            pl.col(name_col).alias("name") if name_col else pl.lit(None).alias("name"),
            pl.col(id_col).alias("id_no") if id_col else pl.lit(None).alias("id_no"),
            pl.col(year_raw_col or year_col).alias("birth_year_raw")
            if (year_raw_col or year_col)
            else pl.lit(None).alias("birth_year_raw"),
        ]
    ).with_columns(
        [
            pl.col("name").cast(pl.Utf8, strict=False).fill_null("").str.strip_chars(),
            pl.col("id_no").cast(pl.Utf8, strict=False).fill_null("").str.strip_chars(),
            pl.col("birth_year_raw")
            .cast(pl.Utf8, strict=False)
            .fill_null("")
            .str.strip_chars()
            .str.replace_all(r"\.0+$", ""),
        ]
    ).with_columns(
        pl.when(pl.col("birth_year_raw").str.contains(r"^\d{4}$"))
        .then(pl.col("birth_year_raw").cast(pl.UInt16, strict=False))
        .otherwise(pl.lit(0, dtype=pl.UInt16))
        .alias("birth_year")
    )


def load_frame(file_path: Path, sheet_name: str) -> pl.DataFrame:
    if file_path.suffix.lower() == ".csv":
        try:
            return pl.read_csv(file_path, try_parse_dates=False, infer_schema_length=None)
        except Exception:
            return pl.read_csv(file_path, encoding="gb18030", try_parse_dates=False, infer_schema_length=None)
    engine = get_settings().import_polars_excel_engine.strip() or "calamine"
    return pl.read_excel(file_path, sheet_name=sheet_name, engine=engine)


def create_stage_table(stage_table_sql: str) -> None:
    clickhouse_command(f"DROP TABLE IF EXISTS {stage_table_sql}")
    clickhouse_command(
        f"""
        CREATE TABLE {stage_table_sql}
        (
            source_file String,
            source_sheet String,
            source_row_no UInt32,
            name String,
            birth_year UInt16,
            birth_year_raw String,
            id_no_cipher String,
            id_no_digest FixedString(64),
            created_by UInt64
        )
        ENGINE = MergeTree
        ORDER BY (source_file, source_sheet, source_row_no)
        """
    )


def stage_table_exists(stage_table_sql: str) -> bool:
    rows = clickhouse_query_rows(f"EXISTS TABLE {stage_table_sql}")
    return bool(rows and int(rows[0]["result"]) == 1)


def count_table_rows(table_sql: str) -> int:
    rows = clickhouse_query_rows(f"SELECT count() AS c FROM {table_sql}")
    return int(rows[0]["c"]) if rows else 0


def stage_source_file(*, file_path: Path, stage_table_sql: str, batch_size: int, created_by: int) -> int:
    settings = get_settings()
    keys = load_keys()
    data_key = keys.data_keys[keys.active_data_key_version]
    total_rows = 0

    for sheet_name in get_sheet_names(file_path):
        df = load_frame(file_path, sheet_name)
        prepared = prepare_stage_frame(df)
        row_count = int(prepared.height)
        if row_count == 0:
            print(f"跳过空 sheet: {file_path.name} / {sheet_name or '(csv)'}", flush=True)
            continue

        names = prepared.get_column("name").to_list()
        raw_ids = prepared.get_column("id_no").to_list()
        birth_years = prepared.get_column("birth_year").to_list()
        birth_year_raws = prepared.get_column("birth_year_raw").to_list()
        row_nos = prepared.get_column("source_row_no").to_list()
        encrypted_ids = encrypt_id_values(data_key, [str(value) for value in raw_ids], workers=max(1, settings.import_encrypt_workers))

        rows: list[dict[str, object]] = []
        for idx in range(row_count):
            rows.append(
                {
                    "source_file": file_path.name,
                    "source_sheet": sheet_name,
                    "source_row_no": int(row_nos[idx]),
                    "name": str(names[idx]),
                    "birth_year": int(birth_years[idx]),
                    "birth_year_raw": str(birth_year_raws[idx]),
                    "id_no_cipher": encrypted_ids[idx],
                    "id_no_digest": fingerprint_id_no(str(raw_ids[idx])),
                    "created_by": int(created_by),
                }
            )
            if len(rows) >= batch_size:
                clickhouse_insert_json_rows(stage_table_sql, rows)
                rows.clear()
        if rows:
            clickhouse_insert_json_rows(stage_table_sql, rows)

        total_rows += row_count
        print(f"已写入 stage: {file_path.name} / {sheet_name or '(csv)'} -> {row_count:,} 行", flush=True)

    return total_rows


def insert_missing_rows(stage_table_sql: str, records_table_sql: str) -> None:
    clickhouse_command(
        f"""
        INSERT INTO {records_table_sql} (id, name, birth_year, birth_year_raw, id_no_cipher, id_no_digest, created_by)
        WITH final_counts AS (
            SELECT
                name,
                birth_year,
                birth_year_raw,
                id_no_digest,
                count() AS existing_count
            FROM {records_table_sql}
            GROUP BY name, birth_year, birth_year_raw, id_no_digest
        )
        SELECT
            cityHash64(source_file, source_sheet, source_row_no) AS id,
            name,
            birth_year,
            birth_year_raw,
            id_no_cipher,
            id_no_digest,
            created_by
        FROM (
            SELECT
                s.*,
                row_number() OVER (
                    PARTITION BY s.name, s.birth_year, s.birth_year_raw, s.id_no_digest
                    ORDER BY s.source_file, s.source_sheet, s.source_row_no
                ) AS rn,
                coalesce(f.existing_count, 0) AS existing_count
            FROM {stage_table_sql} AS s
            LEFT JOIN final_counts AS f
                ON s.name = f.name
               AND s.birth_year = f.birth_year
               AND s.birth_year_raw = f.birth_year_raw
               AND s.id_no_digest = f.id_no_digest
        )
        WHERE rn > existing_count
        """
    )


def main() -> None:
    args = parse_args()
    input_files = args.input_files or DEFAULT_INPUT_FILES
    input_files = [path.expanduser().resolve(strict=True) for path in input_files]

    ensure_clickhouse_record_store()

    settings = get_settings()
    records_table_sql = f"`{settings.clickhouse_database}`.`{settings.clickhouse_records_table}`"
    stage_table_sql = f"`{settings.clickhouse_database}`.`{str(args.stage_table).strip()}`"

    before_rows = count_table_rows(records_table_sql)
    print(f"回填前主表行数: {before_rows:,}", flush=True)

    completed = False
    try:
        if args.use_existing_stage:
            if not stage_table_exists(stage_table_sql):
                raise RuntimeError(f"stage 表不存在，无法复用: {stage_table_sql}")
            staged_rows = count_table_rows(stage_table_sql)
            print(f"复用已有 stage 表，当前行数: {staged_rows:,}", flush=True)
        else:
            create_stage_table(stage_table_sql)
            staged_rows = 0
            for file_path in input_files:
                staged_rows += stage_source_file(
                    file_path=file_path,
                    stage_table_sql=stage_table_sql,
                    batch_size=max(1, int(args.batch_size)),
                    created_by=int(args.created_by),
                )
            print(f"stage 总行数: {staged_rows:,}", flush=True)

        print("开始执行差量回填插入...", flush=True)
        insert_missing_rows(stage_table_sql, records_table_sql)
        after_rows = count_table_rows(records_table_sql)
        print(f"回填后主表行数: {after_rows:,}", flush=True)
        print(f"本次新增行数: {after_rows - before_rows:,}", flush=True)
        completed = True
    finally:
        if completed and not args.keep_stage:
            clickhouse_command(f"DROP TABLE IF EXISTS {stage_table_sql}")
        elif not completed:
            print(f"脚本未完成，保留 stage 表: {stage_table_sql}", flush=True)


if __name__ == "__main__":
    main()
