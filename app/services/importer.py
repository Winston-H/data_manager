from __future__ import annotations

import sqlite3
import threading
import time
from collections.abc import Iterator
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook

from app.core.config import get_settings
from app.core.error_reasons import ErrorReason
from app.core.id_cards import fingerprint_id_no, is_valid_id_no
from app.db.sqlite import open_db_connection
from app.services.audit import write_audit
from app.services.clickhouse_records import existing_id_fingerprints, insert_clickhouse_records

SUPPORTED_IMPORT_EXTENSIONS = frozenset({".xlsx", ".csv"})
IMPORT_SOURCE_DIR = "import_jobs"

_active_import_jobs: set[int] = set()
_active_import_jobs_lock = threading.Lock()
_live_progress: dict[int, dict[str, int | str | None]] = {}
_live_progress_lock = threading.Lock()


@dataclass(frozen=True)
class ImportAuditContext:
    user_id: int
    username: str
    user_role: str
    ip_address: str | None
    trace_id: str | None
    filename: str


@dataclass(frozen=True)
class ImportRuntimeProfile:
    mode: str
    progress_flush_every: int
    live_progress_every: int
    cancel_check_every: int


def supported_import_extensions_text() -> str:
    return ".xlsx/.csv"


def is_supported_import_filename(filename: str) -> bool:
    return Path(filename).suffix.lower() in SUPPORTED_IMPORT_EXTENSIONS


def _normalize_speed_mode(speed_mode: str) -> str:
    return "fast" if str(speed_mode).strip().lower() == "fast" else "normal"


def get_import_runtime_profile() -> ImportRuntimeProfile:
    settings = get_settings()
    mode = _normalize_speed_mode(settings.import_speed_mode)
    if mode == "fast":
        return ImportRuntimeProfile(
            mode="fast",
            progress_flush_every=max(1, settings.import_fast_progress_flush_every),
            live_progress_every=max(1, settings.import_fast_live_progress_every),
            cancel_check_every=max(1, settings.import_fast_cancel_check_every),
        )
    return ImportRuntimeProfile(
        mode="normal",
        progress_flush_every=max(1, settings.import_progress_flush_every),
        live_progress_every=max(1, settings.import_live_progress_every),
        cancel_check_every=max(1, settings.import_cancel_check_every),
    )


def _set_live_progress(
    *,
    job_id: int,
    status: str,
    total_rows: int,
    success_rows: int,
    skipped_rows: int,
    failed_rows: int,
    error_summary: str | None = None,
) -> None:
    with _live_progress_lock:
        _live_progress[job_id] = {
            "status": status,
            "total_rows": total_rows,
            "success_rows": success_rows,
            "skipped_rows": skipped_rows,
            "failed_rows": failed_rows,
            "error_summary": error_summary,
        }


def _clear_live_progress(job_id: int) -> None:
    with _live_progress_lock:
        _live_progress.pop(job_id, None)


def get_live_import_progress(job_id: int) -> dict[str, int | str | None] | None:
    with _live_progress_lock:
        payload = _live_progress.get(job_id)
        return dict(payload) if payload is not None else None


def get_import_source_dir() -> Path:
    data_parent = Path(get_settings().db_path).resolve().parent
    source_dir = data_parent / IMPORT_SOURCE_DIR
    source_dir.mkdir(parents=True, exist_ok=True)
    return source_dir


def get_import_source_path(job_id: int, filename: str | None = None) -> Path:
    suffix = Path(filename or "").suffix.lower() or ".xlsx"
    return get_import_source_dir() / f"job_{job_id}{suffix}"


def create_import_job(conn: sqlite3.Connection, filename: str, file_size_bytes: int, created_by: int) -> int:
    cur = conn.execute(
        """
        INSERT INTO import_jobs(filename, file_size_bytes, status, created_by)
        VALUES (?, ?, 'PENDING', ?)
        """,
        (filename, file_size_bytes, created_by),
    )
    return int(cur.lastrowid)


def get_import_job(conn: sqlite3.Connection, job_id: int) -> sqlite3.Row | None:
    return conn.execute("SELECT * FROM import_jobs WHERE id = ?", (job_id,)).fetchone()


def cancel_import_job(conn: sqlite3.Connection, job_id: int, *, cancelled_by: str) -> sqlite3.Row | None:
    row = get_import_job(conn, job_id)
    if row is None:
        return None
    if row["status"] in {"SUCCESS", "FAILED", "CANCELLED"}:
        return row

    conn.execute(
        """
        UPDATE import_jobs
        SET status = 'CANCELLED',
            finished_at = datetime('now'),
            error_summary = COALESCE(error_summary || ';', '') || ?
        WHERE id = ?
        """,
        (f"cancelled_by={cancelled_by}", job_id),
    )
    conn.commit()
    return get_import_job(conn, job_id)


def list_import_jobs(
    conn: sqlite3.Connection,
    *,
    status: str | None = None,
    created_by: int | None = None,
    filename_contains: str | None = None,
    page: int = 1,
    page_size: int = 20,
) -> tuple[list[sqlite3.Row], int]:
    conditions = []
    params: list[object] = []
    if status:
        conditions.append("status = ?")
        params.append(status)
    if created_by is not None:
        conditions.append("created_by = ?")
        params.append(created_by)
    if filename_contains:
        conditions.append("filename LIKE ?")
        params.append(f"%{filename_contains}%")

    where_sql = f"WHERE {' AND '.join(conditions)}" if conditions else ""
    offset = (page - 1) * page_size
    rows = conn.execute(
        f"""
        SELECT * FROM import_jobs
        {where_sql}
        ORDER BY created_at DESC, id DESC
        LIMIT ? OFFSET ?
        """,
        [*params, page_size, offset],
    ).fetchall()
    total = conn.execute(f"SELECT COUNT(*) AS c FROM import_jobs {where_sql}", params).fetchone()["c"]
    return rows, int(total)


def _flush_progress(
    conn: sqlite3.Connection,
    *,
    job_id: int,
    total_rows: int,
    success_rows: int,
    skipped_rows: int,
    failed_rows: int,
) -> None:
    conn.execute(
        """
        UPDATE import_jobs
        SET total_rows = ?,
            success_rows = ?,
            skipped_rows = ?,
            failed_rows = ?
        WHERE id = ?
        """,
        (total_rows, success_rows, skipped_rows, failed_rows, job_id),
    )
    conn.commit()


def _norm_header(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip().lower().replace("_", "").replace(" ", "")


def _import_polars():
    try:
        import polars as pl
    except ImportError as exc:
        raise RuntimeError("polars is required for import processing") from exc
    return pl


def _resolve_polars_column(columns: list[str], aliases: tuple[str, ...], fallback_idx: int) -> str | None:
    alias_set = {_norm_header(item) for item in aliases}
    for column in columns:
        if _norm_header(column) in alias_set:
            return column
    if 0 <= fallback_idx < len(columns):
        return columns[fallback_idx]
    return None


def _empty_polars_import_frame():
    pl = _import_polars()
    return pl.DataFrame(schema={"name": pl.String, "id_no": pl.String, "birth_year": pl.UInt16})


def _prepare_polars_frame(df) -> tuple[object, int, int]:
    pl = _import_polars()
    columns = [str(col) for col in df.columns]
    if not columns:
        return _empty_polars_import_frame(), 0, 0

    name_col = _resolve_polars_column(columns, ("姓名", "name"), 0)
    id_col = _resolve_polars_column(columns, ("身份证号", "身份证", "证件号", "idno", "id_no"), 1)
    year_col = _resolve_polars_column(columns, ("年份", "year", "年", "birth_year"), 2)

    selected = df.select(
        [
            pl.col(name_col).alias("name") if name_col else pl.lit(None).alias("name"),
            pl.col(id_col).alias("id_no") if id_col else pl.lit(None).alias("id_no"),
            pl.col(year_col).alias("birth_year") if year_col else pl.lit(None).alias("birth_year"),
        ]
    ).with_columns(
        [
            pl.col("name").cast(pl.Utf8, strict=False).fill_null("").str.strip_chars(),
            pl.col("id_no").cast(pl.Utf8, strict=False).fill_null("").str.replace_all(r"\s+", "").str.to_uppercase(),
            pl.col("birth_year")
            .cast(pl.Utf8, strict=False)
            .fill_null("")
            .str.strip_chars()
            .str.replace_all(r"\.0+$", ""),
        ]
    )

    non_blank = selected.filter(
        ~((pl.col("name") == "") & (pl.col("id_no") == "") & (pl.col("birth_year") == ""))
    )
    total_rows = int(non_blank.height)
    if total_rows == 0:
        return _empty_polars_import_frame(), 0, 0

    required = non_blank.filter(
        (pl.col("name") != "") & (pl.col("id_no") != "") & (pl.col("birth_year").str.contains(r"^\d{4}$"))
    )
    deduped = required.unique(subset=["id_no"], keep="first", maintain_order=True)
    valid_mask = [is_valid_id_no(value) for value in deduped.get_column("id_no").to_list()]
    valid = deduped.with_columns(pl.Series(name="_id_valid", values=valid_mask)).filter(pl.col("_id_valid")).drop("_id_valid")
    valid = valid.with_columns(pl.col("birth_year").cast(pl.UInt16, strict=False))
    skipped_rows = total_rows - int(valid.height)
    return valid.select(["name", "id_no", "birth_year"]), total_rows, skipped_rows


def _load_excel_sheet_with_polars(file_path: Path, sheet_name: str):
    pl = _import_polars()
    engine = get_settings().import_polars_excel_engine.strip() or None
    return pl.read_excel(file_path, sheet_name=sheet_name, engine=engine)


def _load_csv_with_polars(file_path: Path):
    pl = _import_polars()
    try:
        return pl.read_csv(file_path, try_parse_dates=False, infer_schema_length=None)
    except Exception:
        return pl.read_csv(file_path, encoding="gb18030", try_parse_dates=False, infer_schema_length=None)


def _iter_polars_frames(file_path: Path) -> Iterator[tuple[object, int, int]]:
    if file_path.suffix.lower() == ".csv":
        yield _prepare_polars_frame(_load_csv_with_polars(file_path))
        return

    wb = load_workbook(file_path, read_only=True, data_only=True)
    try:
        sheet_names = list(wb.sheetnames)
    finally:
        wb.close()

    worker_count = min(max(1, get_settings().import_sheet_read_workers), max(1, len(sheet_names)))
    with ThreadPoolExecutor(max_workers=worker_count, thread_name_prefix="polars-sheet") as executor:
        futures = {
            executor.submit(_load_excel_sheet_with_polars, file_path, sheet_name): sheet_name for sheet_name in sheet_names
        }
        for future in as_completed(futures):
            df = future.result()
            yield _prepare_polars_frame(df)


def run_clickhouse_import_job(
    conn: sqlite3.Connection,
    job_id: int,
    file_path: Path,
    created_by: int,
) -> sqlite3.Row:
    profile = get_import_runtime_profile()
    conn.execute(
        "UPDATE import_jobs SET status='RUNNING', started_at=datetime('now') WHERE id = ?",
        (job_id,),
    )
    conn.commit()
    _set_live_progress(
        job_id=job_id,
        status="RUNNING",
        total_rows=0,
        success_rows=0,
        skipped_rows=0,
        failed_rows=0,
    )

    settings = get_settings()
    t0 = time.perf_counter()
    total_rows = success_rows = skipped_rows = failed_rows = 0
    processed_rows = 0
    cancelled = False

    try:
        for frame, frame_total_rows, frame_skipped_rows in _iter_polars_frames(file_path):
            total_rows += int(frame_total_rows)
            skipped_rows += int(frame_skipped_rows)
            frame_height = int(frame.height)
            if frame_height == 0:
                continue

            names = frame.get_column("name").to_list()
            id_nos = frame.get_column("id_no").to_list()
            birth_years = frame.get_column("birth_year").to_list()

            for idx in range(0, frame_height, max(1, settings.clickhouse_insert_batch_size)):
                batch_names = names[idx : idx + settings.clickhouse_insert_batch_size]
                batch_ids = id_nos[idx : idx + settings.clickhouse_insert_batch_size]
                batch_years = birth_years[idx : idx + settings.clickhouse_insert_batch_size]
                digests = [fingerprint_id_no(id_no) for id_no in batch_ids]
                existing = existing_id_fingerprints(digests)

                final_names: list[str] = []
                final_ids: list[str] = []
                final_years: list[int] = []
                for pos, digest in enumerate(digests):
                    if digest in existing:
                        skipped_rows += 1
                        continue
                    final_names.append(str(batch_names[pos]).strip())
                    final_ids.append(str(batch_ids[pos]).strip())
                    final_years.append(int(batch_years[pos]))

                if final_names:
                    inserted_ids = insert_clickhouse_records(
                        names=final_names,
                        id_nos=final_ids,
                        birth_years=final_years,
                        created_by=created_by,
                    )
                    success_rows += len(inserted_ids)

                processed_rows += len(batch_ids)

                if processed_rows % profile.live_progress_every == 0:
                    _set_live_progress(
                        job_id=job_id,
                        status="RUNNING",
                        total_rows=total_rows,
                        success_rows=success_rows,
                        skipped_rows=skipped_rows,
                        failed_rows=failed_rows,
                    )
                if processed_rows % profile.progress_flush_every == 0:
                    _flush_progress(
                        conn,
                        job_id=job_id,
                        total_rows=total_rows,
                        success_rows=success_rows,
                        skipped_rows=skipped_rows,
                        failed_rows=failed_rows,
                    )
                if processed_rows % profile.cancel_check_every == 0:
                    status_row = conn.execute("SELECT status FROM import_jobs WHERE id = ?", (job_id,)).fetchone()
                    if status_row and status_row["status"] == "CANCELLED":
                        cancelled = True
                        break
            if cancelled:
                break
        elapsed_ms = int((time.perf_counter() - t0) * 1000)
    except Exception as exc:
        elapsed_ms = int((time.perf_counter() - t0) * 1000)
        error_summary = f"import_exception={type(exc).__name__};mode=clickhouse;elapsed_ms={elapsed_ms}"
        conn.execute(
            """
            UPDATE import_jobs
            SET status = 'FAILED',
                total_rows = ?,
                success_rows = ?,
                skipped_rows = ?,
                failed_rows = ?,
                error_summary = ?,
                finished_at = datetime('now')
            WHERE id = ?
            """,
            (total_rows, success_rows, skipped_rows, failed_rows + 1, error_summary, job_id),
        )
        conn.commit()
        _set_live_progress(
            job_id=job_id,
            status="FAILED",
            total_rows=total_rows,
            success_rows=success_rows,
            skipped_rows=skipped_rows,
            failed_rows=failed_rows + 1,
            error_summary=error_summary,
        )
        return get_import_job(conn, job_id)  # type: ignore[return-value]

    if cancelled:
        error_summary = f"cancelled_in_worker=1;mode=clickhouse;elapsed_ms={elapsed_ms}"
        conn.execute(
            """
            UPDATE import_jobs
            SET status = 'CANCELLED',
                total_rows = ?,
                success_rows = ?,
                skipped_rows = ?,
                failed_rows = ?,
                error_summary = COALESCE(error_summary || ';', '') || ?,
                finished_at = datetime('now')
            WHERE id = ?
            """,
            (total_rows, success_rows, skipped_rows, failed_rows, error_summary, job_id),
        )
        conn.commit()
        _set_live_progress(
            job_id=job_id,
            status="CANCELLED",
            total_rows=total_rows,
            success_rows=success_rows,
            skipped_rows=skipped_rows,
            failed_rows=failed_rows,
            error_summary=error_summary,
        )
        return get_import_job(conn, job_id)  # type: ignore[return-value]

    error_summary = f"mode=clickhouse;elapsed_ms={elapsed_ms}"
    conn.execute(
        """
        UPDATE import_jobs
        SET status = 'SUCCESS',
            total_rows = ?,
            success_rows = ?,
            skipped_rows = ?,
            failed_rows = ?,
            error_summary = ?,
            finished_at = datetime('now')
        WHERE id = ?
        """,
        (total_rows, success_rows, skipped_rows, failed_rows, error_summary, job_id),
    )
    conn.commit()
    _set_live_progress(
        job_id=job_id,
        status="SUCCESS",
        total_rows=total_rows,
        success_rows=success_rows,
        skipped_rows=skipped_rows,
        failed_rows=failed_rows,
        error_summary=error_summary,
    )
    return get_import_job(conn, job_id)  # type: ignore[return-value]


def _mark_job_failed(conn: sqlite3.Connection, job_id: int, reason: str) -> None:
    conn.execute(
        """
        UPDATE import_jobs
        SET status = 'FAILED',
            finished_at = datetime('now'),
            error_summary = COALESCE(error_summary || ';', '') || ?
        WHERE id = ?
        """,
        (reason, job_id),
    )


def _run_import_job_worker(
    job_id: int,
    file_path: Path,
    created_by: int,
    audit_ctx: ImportAuditContext,
) -> None:
    conn = open_db_connection()
    try:
        job_row = run_clickhouse_import_job(conn, job_id, file_path, created_by)
        detail = {
            "filename": audit_ctx.filename,
            "total_rows": job_row["total_rows"],
            "success_rows": job_row["success_rows"],
            "failed_rows": job_row["failed_rows"],
            "skipped_rows": job_row["skipped_rows"],
            "status": job_row["status"],
            "async": True,
        }
        if job_row["status"] == "FAILED":
            detail["reason"] = ErrorReason.IMPORT_JOB_FAILED.value
        elif job_row["status"] == "CANCELLED":
            detail["reason"] = ErrorReason.IMPORT_JOB_CANCELLED.value

        write_audit(
            conn,
            user_id=audit_ctx.user_id,
            username=audit_ctx.username,
            user_role=audit_ctx.user_role,
            ip_address=audit_ctx.ip_address,
            action_type="DATA_IMPORT",
            action_result="SUCCESS" if job_row["status"] == "SUCCESS" else "FAILED",
            target_type="IMPORT_JOB",
            target_id=str(job_id),
            detail=detail,
            trace_id=audit_ctx.trace_id,
        )
        conn.commit()
    except Exception as exc:
        _mark_job_failed(conn, job_id, f"worker_exception={type(exc).__name__}")
        write_audit(
            conn,
            user_id=audit_ctx.user_id,
            username=audit_ctx.username,
            user_role=audit_ctx.user_role,
            ip_address=audit_ctx.ip_address,
            action_type="DATA_IMPORT",
            action_result="FAILED",
            target_type="IMPORT_JOB",
            target_id=str(job_id),
            detail={
                "filename": audit_ctx.filename,
                "reason": ErrorReason.IMPORT_WORKER_EXCEPTION.value,
                "error_type": type(exc).__name__,
                "async": True,
            },
            trace_id=audit_ctx.trace_id,
        )
        conn.commit()
    finally:
        with _active_import_jobs_lock:
            _active_import_jobs.discard(job_id)
        _clear_live_progress(job_id)
        conn.close()
        file_path.unlink(missing_ok=True)


def start_import_job_async(
    *,
    job_id: int,
    file_path: Path,
    created_by: int,
    audit_ctx: ImportAuditContext,
) -> bool:
    with _active_import_jobs_lock:
        if job_id in _active_import_jobs:
            return False
        _active_import_jobs.add(job_id)

    thread = threading.Thread(
        target=_run_import_job_worker,
        args=(job_id, file_path, created_by, audit_ctx),
        daemon=True,
        name=f"import-job-{job_id}",
    )
    try:
        thread.start()
    except Exception:
        with _active_import_jobs_lock:
            _active_import_jobs.discard(job_id)
        raise
    return True


def _build_recovery_audit_ctx(conn: sqlite3.Connection, *, created_by: int, filename: str) -> ImportAuditContext:
    user_row = conn.execute("SELECT username, role FROM users WHERE id = ?", (created_by,)).fetchone()
    username = str(user_row["username"]) if user_row else f"user#{created_by}"
    user_role = str(user_row["role"]) if user_row else "UNKNOWN"
    return ImportAuditContext(
        user_id=created_by,
        username=username,
        user_role=user_role,
        ip_address=None,
        trace_id=None,
        filename=filename,
    )


def recover_pending_import_jobs() -> int:
    conn = open_db_connection()
    recovered = 0
    try:
        rows = conn.execute(
            """
            SELECT id, filename, created_by, status
            FROM import_jobs
            WHERE status IN ('PENDING', 'RUNNING')
            ORDER BY id ASC
            """
        ).fetchall()

        for row in rows:
            job_id = int(row["id"])
            created_by = int(row["created_by"])
            filename = str(row["filename"])
            file_path = get_import_source_path(job_id, filename)
            audit_ctx = _build_recovery_audit_ctx(conn, created_by=created_by, filename=filename)

            if not file_path.exists():
                _mark_job_failed(conn, job_id, "source_file_missing=1")
                write_audit(
                    conn,
                    user_id=audit_ctx.user_id,
                    username=audit_ctx.username,
                    user_role=audit_ctx.user_role,
                    ip_address=None,
                    action_type="DATA_IMPORT",
                    action_result="FAILED",
                    target_type="IMPORT_JOB",
                    target_id=str(job_id),
                    detail={
                        "filename": filename,
                        "reason": ErrorReason.IMPORT_WORKER_EXCEPTION.value,
                        "error_type": "SourceFileMissing",
                        "async": True,
                        "recovered": True,
                    },
                    trace_id=None,
                )
                conn.commit()
                continue

            started = start_import_job_async(
                job_id=job_id,
                file_path=file_path,
                created_by=created_by,
                audit_ctx=audit_ctx,
            )
            if started:
                recovered += 1
        return recovered
    finally:
        conn.close()
